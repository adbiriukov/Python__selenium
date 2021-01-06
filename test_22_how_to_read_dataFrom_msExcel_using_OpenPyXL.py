import openpyxl

path = "C:/Users/User/Downloads/111.xlsx"

workbook = openpyxl.load_workbook(path)
sheet = workbook.active  # extract 1 sheet if you have multiple sheets, you must use "get_sheet_by_name(sheet1)"

# how many rows and columns we have. To get number of
rows = sheet.max_row
columns = sheet.max_column
# print(rows) # 3
# print(columns) # 4

for r in range(1, rows+1):
    for c in range(1, columns+1):
        print(sheet.cell(row=r, column=c).value, end="    ")
    print()